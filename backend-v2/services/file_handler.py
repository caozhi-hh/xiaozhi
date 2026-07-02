"""
文件处理 — PDF/Word/Excel 文字提取 + 图片 base64 编码

支持的文件类型：
  PDF   → 提取文字内容
  Word  → 提取文字内容（.docx）
  Excel → 提取为表格文本（.xlsx）
  图片（PNG/JPG/JPEG/GIF/WEBP）→ base64 编码，走视觉模型
"""
import base64
from io import BytesIO

import fitz  # PyMuPDF

IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}
DOCUMENT_EXTENSIONS = {"pdf", "docx", "xlsx", "txt", "md"}


def is_image(filename: str) -> bool:
    return _ext(filename) in IMAGE_EXTENSIONS


def is_pdf(filename: str) -> bool:
    return _ext(filename) == "pdf"


def is_docx(filename: str) -> bool:
    return _ext(filename) == "docx"


def is_xlsx(filename: str) -> bool:
    return _ext(filename) == "xlsx"


def is_document(filename: str) -> bool:
    return _ext(filename) in DOCUMENT_EXTENSIONS


def _ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def extract_pdf_text(file_bytes: bytes) -> str:
    """从 PDF 中提取全部文字"""
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()
    return text.strip()


def extract_docx_text(file_bytes: bytes) -> str:
    """从 Word (.docx) 中提取全部文字"""
    from docx import Document as DocxDocument
    doc = DocxDocument(BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip()).strip()


def extract_xlsx_text(file_bytes: bytes) -> str:
    """从 Excel (.xlsx) 中提取为表格文本"""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(file_bytes), read_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"## {sheet}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append(" | ".join(cells))
        lines.append("")
    wb.close()
    return "\n".join(lines).strip()


def extract_text(file_bytes: bytes, filename: str) -> str:
    """根据文件类型自动提取文字"""
    if is_pdf(filename):
        return extract_pdf_text(file_bytes)
    elif is_docx(filename):
        return extract_docx_text(file_bytes)
    elif is_xlsx(filename):
        return extract_xlsx_text(file_bytes)
    elif _ext(filename) in ("txt", "md"):
        return file_bytes.decode("utf-8", errors="ignore").strip()
    return ""


def image_to_base64(file_bytes: bytes, filename: str) -> str:
    """将图片转为 base64 data URL（自动压缩大图）"""
    from PIL import Image

    img = Image.open(BytesIO(file_bytes))

    max_size = 1024
    if max(img.size) > max_size:
        ratio = max_size / max(img.size)
        new_size = (int(img.width * ratio), int(img.height * ratio))
        img = img.resize(new_size)

    buf = BytesIO()
    img = img.convert("RGB")
    img.save(buf, format="JPEG", quality=85)
    b64 = base64.b64encode(buf.getvalue()).decode()
    return f"data:image/jpeg;base64,{b64}"
